from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from mysql.connector import Error
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.api.deps import require_roles
from app.db.connection import get_connection

router = APIRouter(prefix="/exports", tags=["exports"])


def _build_xlsx(headers: list[str], rows: list[tuple], sheet_name: str) -> BytesIO:
    buffer = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center")

    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        worksheet.append(["" if value is None else value for value in row])

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 3, 40)

    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/products")
def export_products(
    warehouse_id: int | None = Query(default=None),
    _: dict = Depends(require_roles("Admin", "Manager", "Clerk")),
):
    conn = get_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")

    try:
        cur = conn.cursor()
        query = """
            SELECT p.product_id, p.name, p.description, p.current_stock,
                   p.unit_price, p.unit,
                   COALESCE(w.name, '') AS warehouse,
                   COALESCE(c.name, '') AS category,
                   p.low_stock_threshold, p.expiry_date, p.expiry_status,
                   p.manufactured_date, p.batch_number
            FROM products p
            LEFT JOIN warehouses w ON p.warehouse_id = w.warehouse_id
            LEFT JOIN categories c ON p.category_id = c.category_id
        """
        params: list[object] = []
        if warehouse_id:
            query += " WHERE p.warehouse_id = %s"
            params.append(warehouse_id)
        query += " ORDER BY p.product_id"
        cur.execute(query, tuple(params)) if params else cur.execute(query)
        rows = cur.fetchall()
        cur.close()

        headers = [
            "ID",
            "Name",
            "Description",
            "Stock",
            "Unit Price (PHP)",
            "Unit",
            "Warehouse",
            "Category",
            "Low Stock Threshold",
            "Expiry Date",
            "Expiry Status",
            "Manufactured Date",
            "Batch Number",
        ]
        buffer = _build_xlsx(headers, rows, "Products")
        filename = "products.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Error:
        raise HTTPException(status_code=500, detail="Failed to export products")
    finally:
        conn.close()


@router.get("/transactions")
def export_transactions(
    warehouse_id: int | None = Query(default=None),
    search: str | None = Query(default=None),
    txn_type: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    _: dict = Depends(require_roles("Admin", "Manager")),
):
    conn = get_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")

    try:
        cur = conn.cursor()
        query = """
            SELECT t.transaction_id, p.name, t.type, t.quantity,
                   COALESCE(t.unit_cost, p.unit_price) AS unit_cost,
                   ROUND(COALESCE(t.unit_cost, p.unit_price) * t.quantity, 2) AS total_cost,
                   DATE_FORMAT(t.transaction_date, %s),
                   t.remarks,
                   COALESCE(u.username, t.user_id),
                   COALESCE(w.name, ''),
                   COALESCE(c.name, ''),
                   COALESCE(p.batch_number, '')
            FROM transactions t
            JOIN products p ON t.product_id = p.product_id
            LEFT JOIN users u ON t.user_id = u.user_id
            LEFT JOIN warehouses w ON t.warehouse_id = w.warehouse_id
            LEFT JOIN categories c ON p.category_id = c.category_id
            WHERE 1=1
        """
        params: list[object] = ["%Y-%m-%d %H:%i"]

        if warehouse_id:
            query += " AND t.warehouse_id = %s"
            params.append(warehouse_id)
        if search:
            query += " AND p.name LIKE %s"
            params.append(f"%{search}%")
        if txn_type and txn_type != "All":
            query += " AND t.type = %s"
            params.append(txn_type)
        if date_from:
            query += " AND DATE(t.transaction_date) >= %s"
            params.append(date_from)
        if date_to:
            query += " AND DATE(t.transaction_date) <= %s"
            params.append(date_to)

        query += " ORDER BY t.transaction_date DESC"
        cur.execute(query, tuple(params))
        rows = cur.fetchall()
        cur.close()

        headers = [
            "ID",
            "Product",
            "Type",
            "Quantity",
            "Unit Cost (PHP)",
            "Total Cost (PHP)",
            "Date",
            "Remarks",
            "User",
            "Warehouse",
            "Category",
            "Batch",
        ]
        buffer = _build_xlsx(headers, rows, "Transactions")
        filename = "transactions.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Error:
        raise HTTPException(status_code=500, detail="Failed to export transactions")
    finally:
        conn.close()
