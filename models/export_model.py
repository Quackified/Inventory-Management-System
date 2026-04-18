"""
export_model.py — Export products and transactions to CSV and Excel.
"""

import csv
from database import get_connection, close_connection
from mysql.connector import Error


def _get_product_rows(warehouse_id=None):
    """Fetch product data for export."""
    conn = get_connection()
    if not conn:
        return [], []
    try:
        cur = conn.cursor()
        query = """
            SELECT p.product_id, p.name, p.description, p.current_stock,
                   p.unit,
                   COALESCE(w.name, '') AS warehouse,
                   COALESCE(c.name, '') AS category,
                   p.low_stock_threshold, p.expiry_date, p.expiry_status,
                   p.manufactured_date, p.batch_number
            FROM products p
            LEFT JOIN warehouses w ON p.warehouse_id = w.warehouse_id
            LEFT JOIN categories c ON p.category_id = c.category_id
        """
        params = []
        if warehouse_id:
            query += " WHERE p.warehouse_id = %s"
            params.append(warehouse_id)
        query += " ORDER BY p.product_id"

        if params:
            cur.execute(query, tuple(params))
        else:
            cur.execute(query)
        rows = cur.fetchall()
        headers = ["ID", "Name", "Description", "Stock", "Unit",
                    "Warehouse", "Category", "Low Stock Threshold",
                    "Expiry Date", "Expiry Status",
                    "Manufactured Date", "Batch Number"]
        cur.close()
        return headers, rows
    except Error as e:
        print(f"[ExportModel] Product export error: {e}")
        return [], []
    finally:
        close_connection(conn)


def _get_transaction_rows(warehouse_id=None, search_term=None,
                           txn_type=None, date_from=None, date_to=None):
    """Fetch transaction data for export."""
    conn = get_connection()
    if not conn:
        return [], []
    try:
        cur = conn.cursor()
        query = """
            SELECT t.transaction_id, p.name, t.type, t.quantity,
                   DATE_FORMAT(t.transaction_date, %s),
                   t.remarks,
                   COALESCE(u.username, t.user_id),
                   COALESCE(w.name, ''),
                   COALESCE(c.name, ''),
                   COALESCE(p.batch_number, '')
            FROM transactions t
            JOIN products p ON t.product_id = p.product_id
            LEFT JOIN users u ON t.user_id = u.user_id
            LEFT JOIN warehouses w ON t.warehouse_id = w.warehouse_id
            LEFT JOIN categories c ON p.category_id = c.category_id
            WHERE 1=1
        """
        params = ["%Y-%m-%d %H:%i"]

        if warehouse_id:
            query += " AND t.warehouse_id = %s"
            params.append(warehouse_id)
        if search_term:
            query += " AND p.name LIKE %s"
            params.append(f"%{search_term}%")
        if txn_type and txn_type != "All":
            query += " AND t.type = %s"
            params.append(txn_type)
        if date_from:
            query += " AND DATE(t.transaction_date) >= %s"
            params.append(date_from)
        if date_to:
            query += " AND DATE(t.transaction_date) <= %s"
            params.append(date_to)

        query += " ORDER BY t.transaction_date DESC"
        cur.execute(query, tuple(params))
        rows = cur.fetchall()
        headers = ["ID", "Product", "Type", "Quantity", "Date",
                    "Remarks", "User", "Warehouse", "Category", "Batch"]
        cur.close()
        return headers, rows
    except Error as e:
        print(f"[ExportModel] Transaction export error: {e}")
        return [], []
    finally:
        close_connection(conn)


# ── CSV Export ───────────────────────────────────────────────
def export_products_csv(filepath, warehouse_id=None):
    headers, rows = _get_product_rows(warehouse_id)
    if not headers:
        return False, "No data to export."
    try:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        return True, f"Exported {len(rows)} products to CSV."
    except Exception as e:
        return False, str(e)


def export_transactions_csv(filepath, **filters):
    headers, rows = _get_transaction_rows(**filters)
    if not headers:
        return False, "No data to export."
    try:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        return True, f"Exported {len(rows)} transactions to CSV."
    except Exception as e:
        return False, str(e)


# ── Excel Export (xlsx) ──────────────────────────────────────
def export_products_xlsx(filepath, warehouse_id=None):
    headers, rows = _get_product_rows(warehouse_id)
    if not headers:
        return False, "No data to export."
    return _write_xlsx(filepath, "Products", headers, rows)


def export_transactions_xlsx(filepath, **filters):
    headers, rows = _get_transaction_rows(**filters)
    if not headers:
        return False, "No data to export."
    return _write_xlsx(filepath, "Transactions", headers, rows)


def _write_xlsx(filepath, sheet_name, headers, rows):
    """Write data to an Excel file using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False, ("openpyxl is not installed.\n"
                       "Run: pip install openpyxl")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", fill_type="solid")
        header_align = Alignment(horizontal="center")

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row in rows:
            ws.append([str(v) if v is not None else "" for v in row])

        # Auto-fit column widths (approximate)
        for i, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        wb.save(filepath)
        return True, f"Exported {len(rows)} {sheet_name.lower()} to Excel."
    except Exception as e:
        return False, str(e)
